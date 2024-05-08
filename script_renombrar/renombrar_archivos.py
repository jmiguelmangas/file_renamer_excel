import os
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox

def rename_files():
    excel_file_path = excel_file_entry.get()
    folder_path = folder_entry.get()

    try:
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active
    except Exception as e:
        messagebox.showerror("Error", f"Error al cargar el archivo de Excel: {e}")
        return

    num_rows_excel = sheet.max_row

    files = os.listdir(folder_path)
    num_files = len(files)

    if num_files != num_rows_excel:
        messagebox.showerror("Error", "El número de archivos en la carpeta no coincide con el número de filas menos una en el archivo Excel.")
        return

    for i in range(1, num_rows_excel + 1):
        new_filename = sheet.cell(row=i, column=2).value
        if new_filename:
            old_file_path = os.path.join(folder_path, f"{i:03d}.wav")
            new_file_path = os.path.join(folder_path, f"{new_filename}.wav")
            os.rename(old_file_path, new_file_path)


    messagebox.showinfo("Éxito", "Proceso completado con éxito.")

def browse_excel_file():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, filename)

def browse_folder():
    foldername = filedialog.askdirectory()
    folder_entry.delete(0, tk.END)
    folder_entry.insert(0, foldername)

# Crear la ventana principal
root = tk.Tk()
root.title("Renombrar archivos desde Excel")

# Crear los widgets
excel_file_label = tk.Label(root, text="Archivo de Excel:")
excel_file_entry = tk.Entry(root, width=50)
excel_file_button = tk.Button(root, text="Buscar", command=browse_excel_file)

folder_label = tk.Label(root, text="Carpeta de archivos:")
folder_entry = tk.Entry(root, width=50)
folder_button = tk.Button(root, text="Buscar", command=browse_folder)

rename_button = tk.Button(root, text="Renombrar archivos", command=rename_files)

# Ubicar los widgets en la ventana
excel_file_label.grid(row=0, column=0, padx=10, pady=5, sticky="e")
excel_file_entry.grid(row=0, column=1, padx=10, pady=5)
excel_file_button.grid(row=0, column=2, padx=10, pady=5)

folder_label.grid(row=1, column=0, padx=10, pady=5, sticky="e")
folder_entry.grid(row=1, column=1, padx=10, pady=5)
folder_button.grid(row=1, column=2, padx=10, pady=5)

rename_button.grid(row=2, column=0, columnspan=3, padx=10, pady=10)

# Ejecutar la aplicación
root.mainloop()
